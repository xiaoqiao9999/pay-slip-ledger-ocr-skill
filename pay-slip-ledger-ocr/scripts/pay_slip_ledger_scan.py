# -*- coding: utf-8 -*-
"""
付款回单 OCR → 字段提取 → 台账生成 + 重命名归档
================================================
输入   : 工作区「输入」目录下的付款回单 PNG（财务微信图片）
处理   : 调用本地 NPU OCR (run.ps1) 逐张识别 → 正则提取 12 个台账字段
输出   : 付款回单台账.xlsx（12 列，追加式；幂等去重依据 = 会计流水号）
归档   : 识别成功且字段完整的 PNG，重命名为
         「收款人名称_金额两位小数_摘要_记账日期.png」后移入
         已完成目录下以「收款人名称」命名的子文件夹（按供应商分类）
         重名自动加 _2/_3 后缀；字段缺失的留在「输入」并标"待人工核对"

用法   :
  python 扫描付款回单生成台账.py                 # 正式运行
  python 扫描付款回单生成台账.py --dry-run       # 只预览不写盘不移动
  python 扫描付款回单生成台账.py --input <dir> --output <xlsx> --ocr-script <ps1>
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# 路径配置（可被命令行参数覆盖）
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "输入")
DONE_DIR = os.path.join(BASE_DIR, "已完成")
OUTPUT_XLSX = os.path.join(BASE_DIR, "付款回单台账.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "扫描日志.txt")


def _find_ocr_script() -> str:
    """探测本地 NPU OCR 引擎 (run.ps1)：
    1) 环境变量 PAYSLIP_OCR_SCRIPT 优先；2) 本机常见安装路径；3) 兜底默认值。
    其他机器部署时可通过环境变量或 --ocr-script 显式指定。"""
    env = os.environ.get("PAYSLIP_OCR_SCRIPT")
    if env and os.path.exists(env):
        return env
    cands = [
        r"c:\Users\JQZH\.trae-cn\skills\local-ocr-npu\scripts\run.ps1",
        r"c:\Users\JQZH\.workbuddy\skills\local-ocr-npu\scripts\run.ps1",
    ]
    for c in cands:
        if os.path.exists(c):
            return c
    return cands[0]


OCR_SCRIPT = _find_ocr_script()

# ============================================================
# 日志（控制台 + 文件双路）
# ============================================================
log = logging.getLogger("pay_slip_scan")


def setup_logging(log_path: str):
    log.setLevel(logging.INFO)
    if log.handlers:
        log.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    log.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    log.addHandler(ch)


# ============================================================
# OCR 调用（复用 一键生成_优化版.py 的成熟范式）
# ============================================================
def ocr_one(img_path: str, index: int, ocr_script: str, tmp_root: str, timeout=300) -> list[str]:
    """使用 NPU OCR 识别单张图片（强制 CPU 模式）。
    ppocr.exe 不支持中文路径/文件名，先复制到纯英文临时文件。"""
    os.makedirs(tmp_root, exist_ok=True)
    tmp_img = os.path.join(tmp_root, "img_%04d.jpg" % index)
    try:
        shutil.copy2(img_path, tmp_img)
    except Exception as e:
        log.warning("复制临时文件失败 %s: %s", img_path, e)
        return []
    try:
        r = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", ocr_script, tmp_img, "-Device", "cpu"],
            capture_output=True, timeout=timeout,
        )
        raw = r.stdout
        try:
            out = raw.decode("utf-8")
        except UnicodeDecodeError:
            out = raw.decode("gbk", errors="replace")
        if r.returncode != 0:
            err = [l for l in out.splitlines() if "[ERROR]" in l]
            msg = "; ".join(err) if err else ("exit code %d" % r.returncode)
            log.warning("OCR失败 %s: %s", os.path.basename(img_path), msg)
            return []
        lines: list[str] = []
        inside = False
        for line in out.splitlines():
            s = line.strip()
            if s.startswith("---") and len(s) >= 40:
                inside = not inside
                continue
            if inside and s:
                lines.append(s)
        return lines
    except subprocess.TimeoutExpired:
        log.warning("OCR超时(>%ss) %s", timeout, os.path.basename(img_path))
        return []
    except Exception as e:
        log.warning("OCR异常 %s: %s", os.path.basename(img_path), e)
        return []
    finally:
        if os.path.exists(tmp_img):
            try:
                os.remove(tmp_img)
            except OSError:
                pass


# ============================================================
# 字段提取正则（table-driven，加字段只需在此加一行）
# ============================================================
FIELDS = {
    "记账日期":  r"记账日期\s*[:：]?\s*(\d{4})\s*(\d{2})\s*(\d{2})",
    "付款人名称": r"付款人名称\s*[:：]?\s*([^\n]+)",
    "收款人名称": r"收款人名称\s*[:：]?\s*([^\n]+)",
    "收款人账号": r"收款人账号\s*[:：]?\s*(\d[\d\s]*)",
    "开户行":    r"开户行名称\s*[:：]?\s*([^\n]+)",
    "币种":     r"币种\s*[:：]?\s*([A-Za-z]{3})",
    "金额":     r"(?<!大写)金额\s*[:：]?\s*([\d,，][\d,，\.]*)",
    "金额大写":  r"金额大写\s*[:：]?\s*([^\n]+)",
    "摘要":     r"摘要\s*[:：]?\s*([^\n]+)",
    "会计流水号": r"会计流水号\s*[:：]?\s*([A-Za-z0-9]+)",
    "批次元号":  r"批次(?:元|原)?号\s*[:：]?\s*(\d+)",
}

# 台账表头列顺序（第 1 列序号；无源文件列，去重靠第 11 列会计流水号）
HEADERS = ["序号", "记账日期", "付款人名称", "收款人名称", "收款人账号",
           "开户行", "币种", "金额", "金额大写", "摘要", "会计流水号", "批次元号"]
MONEY_COL = 8    # 金额列（1 起）
ID_COL = 11      # 会计流水号列（去重键）

# 必须项（缺任一 → 行标"待人工核对"、文件留输入目录）；其余字段均为可选项
REQUIRED_FIELDS = ("收款人名称", "金额", "摘要")


def normalize_colon(text: str) -> str:
    return text.replace("：", ":").replace("（", "(").replace("）", ")")


def extract_fields(lines: list[str]) -> dict:
    """从 OCR 行文本提取台账字段。缺失字段填 '待人工核对'。"""
    text = normalize_colon("\n".join(lines))
    out: dict[str, object] = {}
    missing: list[str] = []
    for key, pat in FIELDS.items():
        if key == "开户行":
            banks = re.findall(pat, text)
            out[key] = banks[1].strip() if len(banks) >= 2 else (banks[0].strip() if banks else "")
        elif key == "记账日期":
            m = re.search(pat, text)
            out[key] = ("%s-%s-%s" % (m.group(1), m.group(2), m.group(3))) if m else ""
        elif key == "金额":
            m = re.search(pat, text)
            if m:
                digits = re.sub(r"[,\s，]", "", m.group(1))
                # 仅保留最多两位小数的数字（防 OCR 粘入多余点号）
                mm = re.match(r"^(\d+)(?:\.(\d{1,2}))?", digits)
                out[key] = (mm.group(1) + ("." + mm.group(2) if mm.group(2) else "")) if mm else ""
            else:
                out[key] = ""
        else:
            m = re.search(pat, text)
            val = m.group(1).strip() if m else ""
            if key == "收款人账号":
                val = re.sub(r"\s+", "", val)
            out[key] = val
    # 必须项缺失 → 标"待人工核对"；可选字段（记账日期/流水号/批次号等）缺失 → 留空
    for k in out:
        if not out[k]:
            out[k] = "待人工核对" if k in REQUIRED_FIELDS else ""
    missing = [k for k in REQUIRED_FIELDS if out[k] == "待人工核对"]
    out["__missing__"] = missing
    return out


# ============================================================
# 归档重命名：收款人名称_金额(两位小数)_摘要_记账日期
# ============================================================
ILLEGAL_CHARS = re.compile(r'[\\/:*?"<>|\r\n\t]')


def sanitize(s: str) -> str:
    s = ILLEGAL_CHARS.sub("_", str(s))
    s = re.sub(r"\s+", "", s)
    return s[:40] or "待人工核对"


def build_new_name(fields: dict) -> str:
    """收款人_金额两位小数_摘要_记账日期.png；记账日期(可选)缺失时省略日期段。"""
    payee = sanitize(fields["收款人名称"])
    memo = sanitize(fields["摘要"])
    amt_raw = str(fields["金额"]).replace("待人工核对", "0") or "0"
    try:
        amt = "%.2f" % Decimal(amt_raw)   # 两位小数：13920.00 / 3927.17
    except (InvalidOperation, ValueError):
        amt = "0.00"
    parts = [payee, amt, memo]
    date = sanitize(fields["记账日期"])
    if date and date != "待人工核对":
        parts.append(date)
    return "_".join(parts) + ".png"


def unique_dest(done_dir: str, new_name: str) -> str:
    """重名自动加 _2/_3 后缀，返回可用目标路径。"""
    stem, ext = os.path.splitext(new_name)
    cand = os.path.join(done_dir, new_name)
    n = 2
    while os.path.exists(cand):
        cand = os.path.join(done_dir, "%s_%d%s" % (stem, n, ext))
        n += 1
    return cand


# ============================================================
# Excel 台账：表头 / 追加 / 去重 / 合计
# ============================================================
def ensure_headers(xlsx_path: str):
    wb = Workbook() if not os.path.exists(xlsx_path) else load_workbook(xlsx_path)
    ws = wb.active
    if ws.max_row == 0 or ws.cell(1, 1).value != HEADERS[0]:
        # 已存在旧合计行(空行起始)时先清空重建
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(1, c)
            cell.value = HEADERS[c - 1]
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        # 列宽
        widths = [6, 12, 30, 32, 22, 34, 8, 14, 34, 26, 22, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
    wb.save(xlsx_path)


def load_seen_ids(xlsx_path: str) -> set[str]:
    """读取台账已记录的会计流水号集合（用于幂等去重）。"""
    if not os.path.exists(xlsx_path):
        return set()
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        seen = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, ID_COL).value
            if v and str(v).strip() not in ("", "待人工核对"):
                seen.add(str(v).strip())
        return seen
    except Exception as e:
        log.warning("读取台账去重失败: %s", e)
        return set()


def append_and_total(xlsx_path: str, row_data: list):
    """单次加载：删除旧合计 → 追加新数据行 → 重建合计行（Decimal 求和）。

    避免"追加与合计分两次读写"导致的孤儿行/行错位问题。"""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 1) 删除已有合计行（可能位于任意位置，一律清掉）
    del_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == "合计"]
    for r in reversed(del_rows):
        ws.delete_rows(r)

    # 2) 真实数据末行（按第 1 列序号为数字判断）
    last_data = 1
    for r in range(2, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        if v1 is not None and isinstance(v1, (int, float)):
            last_data = r

    # 3) 追加新数据行并写序号（先重排已有序号保证连续）
    seq = 1
    for r in range(2, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        if v1 is not None and isinstance(v1, (int, float)):
            ws.cell(r, 1).value = seq
            seq += 1
    row_data[0] = seq
    ws.append(row_data)
    new_r = ws.max_row
    ws.cell(new_r, MONEY_COL).number_format = "#,##0.00"

    # 4) 重建合计行
    total = Decimal(0)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "合计":
            continue
        v = ws.cell(r, MONEY_COL).value
        if v is None or str(v).strip() in ("", "待人工核对"):
            continue
        try:
            total += Decimal(str(v))
        except (InvalidOperation, ValueError):
            continue
    total_row = ws.max_row + 1
    ws.cell(total_row, 1).value = "合计"
    ws.cell(total_row, MONEY_COL).value = float(total)
    ws.cell(total_row, MONEY_COL).number_format = "#,##0.00"
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(total_row, c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
    wb.save(xlsx_path)


# ============================================================
# 主流程
# ============================================================
def main():
    ap = argparse.ArgumentParser(description="付款回单 PNG OCR → 台账 + 重命名归档")
    ap.add_argument("--input", default=INPUT_DIR, help="输入目录（默认: 转账回单/输入）")
    ap.add_argument("--output", default=OUTPUT_XLSX, help="台账 xlsx 路径")
    ap.add_argument("--done", default=DONE_DIR, help="已完成归档目录")
    ap.add_argument("--ocr-script", default=OCR_SCRIPT, help="本地 OCR run.ps1 路径")
    ap.add_argument("--dry-run", action="store_true", help="只预览，不写台账、不移动文件")
    ap.add_argument("--limit", type=int, default=0, help="只处理前 N 张（测试用，0=全部）")
    args = ap.parse_args()

    input_dir = os.path.abspath(args.input)
    done_dir = os.path.abspath(args.done)
    xlsx_path = os.path.abspath(args.output)
    log_path = os.path.join(os.path.dirname(xlsx_path), "扫描日志.txt")

    setup_logging(log_path)
    if args.dry_run:
        log.info("=" * 60)
        log.info("DRY-RUN 模式：不写台账、不移动文件，仅预览解析结果")
        log.info("=" * 60)

    if not os.path.isdir(input_dir):
        log.error("输入目录不存在: %s", input_dir)
        sys.exit(1)
    if not os.path.exists(args.ocr_script):
        log.error("OCR 脚本不存在: %s", args.ocr_script)
        sys.exit(1)
    os.makedirs(done_dir, exist_ok=True)

    png_files = sorted(f for f in os.listdir(input_dir)
                       if f.lower().endswith((".png", ".jpg", ".jpeg")))
    if args.limit and args.limit > 0:
        png_files = png_files[:args.limit]
    if not png_files:
        log.info("输入目录没有待处理的图片: %s", input_dir)
        return

    seen = set() if args.dry_run else load_seen_ids(xlsx_path)
    if not args.dry_run and os.path.exists(xlsx_path):
        ensure_headers(xlsx_path)

    log.info("待处理 %d 张图片（台账已有 %d 条记录）", len(png_files), len(seen))
    tmp_root = os.path.join(tempfile.gettempdir(), "pay_slip_ocr_tmp")
    stats = {"ok": 0, "missing": 0, "fail": 0, "skip": 0}

    for idx, fname in enumerate(png_files):
        src = os.path.join(input_dir, fname)
        log.info("OCR(%d/%d): %s", idx + 1, len(png_files), fname)
        lines = ocr_one(src, idx, args.ocr_script, tmp_root)
        if not lines:
            log.warning("  → OCR 未返回文本，文件留在输入目录待人工处理")
            stats["fail"] += 1
            continue
        fields = extract_fields(lines)
        missing = fields["__missing__"]
        # 幂等去重：会计流水号已在台账 → 跳过（不重复追加）
        sid = str(fields["会计流水号"]).strip()
        if sid not in ("待人工核对", "") and sid in seen:
            log.info("跳过(会计流水号已在台账): %s [%s]", fname, sid)
            stats["skip"] += 1
            continue
        new_name = build_new_name(fields)

        if args.dry_run:
            ok_mark = "✓完整" if not missing else "✗缺%s" % ",".join(missing)
            log.info("  [%s] %s", ok_mark, fname)
            log.info("        收款人=%s | 金额=%s | 摘要=%s | 日期=%s | 流水号=%s",
                     fields["收款人名称"], fields["金额"], fields["摘要"], fields["记账日期"], sid)
            log.info("        新文件名: %s", new_name)
            if missing:
                stats["missing"] += 1
            else:
                stats["ok"] += 1
            continue

        # ===== 正式写入 =====
        ensure_headers(xlsx_path)
        c = lambda v: None if v == "" else v   # 可选字段缺失 → 空单元格
        row_data = [0,
                    c(fields["记账日期"]), fields["付款人名称"], fields["收款人名称"],
                    c(fields["收款人账号"]), c(fields["开户行"]), c(fields["币种"]),
                    fields["金额"], c(fields["金额大写"]), fields["摘要"],
                    c(fields["会计流水号"]), c(fields["批次元号"])]
        append_and_total(xlsx_path, row_data)

        if missing:
            log.warning("  → 已写入台账，但缺必须字段(%s) → 文件留输入目录待人工核对",
                        ",".join(missing))
            stats["missing"] += 1
        else:
            # 归档：已完成\<收款人>\ 子目录（按供应商分类）
            payee_dir = sanitize(fields["收款人名称"])
            dest_dir = os.path.join(done_dir, payee_dir)
            os.makedirs(dest_dir, exist_ok=True)
            dest = unique_dest(dest_dir, new_name)
            try:
                shutil.move(src, dest)
                log.info("  ✓ 归档: %s → %s\\%s", fname, payee_dir, os.path.basename(dest))
                stats["ok"] += 1
            except Exception as e:
                log.warning("  → 台账已写入但移动失败: %s", e)
                stats["missing"] += 1

    log.info("=" * 60)
    log.info("处理完成: 完整归档=%d | 待人工核对=%d | OCR失败=%d | 跳过=%d",
             stats["ok"], stats["missing"], stats["fail"], stats["skip"])
    if args.dry_run:
        log.info("(DRY-RUN 未写入任何文件)")
    else:
        log.info("台账: %s", xlsx_path)
        log.info("归档目录: %s", done_dir)
        log.info("日志: %s", log_path)


if __name__ == "__main__":
    main()
